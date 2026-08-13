from django.shortcuts import render, redirect
from accounts.models import Store
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
from .models import *
from PIL import Image
from django.core.files.base import ContentFile
from io import BytesIO
from .utils import delete_old_reports
from django.utils.dateparse import parse_date
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from django.http import HttpResponse
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from io import BytesIO
from PIL import Image
from django.core.files.uploadedfile import InMemoryUploadedFile
import sys
import os
import uuid
import cloudinary.uploader
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login
from .models import Audit, AuditIssue
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Exists, OuterRef  
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import *
from django.db.models import Count
from django.db.models import Avg, Count
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db.models import Avg, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from itertools import groupby

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import requests
from io import BytesIO
import traceback

def dashboard(request):
    try:
        #delete_old_reports()

        all_slots = TimeSlot.objects.all().order_by("start_time")

        if "store_id" not in request.session:
            return redirect("/")

        store = Store.objects.get(
            id=request.session["store_id"]
        )

        current_slot = get_current_slot()

        items = ChecklistItem.objects.filter(
            slot=current_slot
        )

        now = timezone.localtime()
        today = now.date()

        if now.hour < 3:
            today = today - timedelta(days=1)

        completed_ids = Report.objects.filter(
            store=store,
            report_date=today
        ).values_list(
            "item_id",
            flat=True
        )

        return render(
            request,
            "checklist/dashboard.html",
            {
                "store": store,
                "items": items,
                "slot": current_slot,
                "all_slots": all_slots,
                "completed_ids": list(completed_ids)
            }
        )

    except Exception:
        print(traceback.format_exc())
        raise

def get_current_slot():

    now = datetime.now().time()

    slots = TimeSlot.objects.all()

    for slot in slots:

        if slot.start_time <= slot.end_time:

            if slot.start_time <= now <= slot.end_time:
                return slot

        else:

            if now >= slot.start_time or now <= slot.end_time:
                return slot

    return None

from PIL import Image, UnidentifiedImageError

def compress_image(image_file):
    try:
        img = Image.open(image_file)

        print("FORMAT =", img.format)

        img = img.convert("RGB")
        img.thumbnail((1280, 1280))

        output = BytesIO()

        img.save(
            output,
            format="JPEG",
            quality=75
        )

        output.seek(0)

        return ContentFile(output.read())

    except UnidentifiedImageError:
        raise Exception(
            f"Không đọc được ảnh: {image_file.name}"
        )

def upload_report(request, item_id):
    if "store_id" not in request.session:
        return redirect("/")

    store = Store.objects.get(id=request.session["store_id"])
    item = ChecklistItem.objects.get(id=item_id)

    if request.method == "POST":
        try:
            print("=== START UPLOAD ===")

            if "image" not in request.FILES:
                print("NO IMAGE")
                return redirect("/dashboard/")

            image_file = request.FILES["image"]

            print("FILE NAME:", image_file.name)
            print("CONTENT TYPE:", image_file.content_type)

            now = timezone.localtime()
            report_date = now.date()

            if now.hour < 3:
                report_date = report_date - timedelta(days=1)

            Report.objects.filter(
                store=store,
                item=item,
                report_date=report_date
            ).delete()

            random_public_id = f"mixue_{uuid.uuid4().hex[:10]}"

            print("COMPRESSING...")
            compressed_image = compress_image(image_file)

            print("UPLOADING TO CLOUDINARY...")
            upload_result = cloudinary.uploader.upload(
                compressed_image,
                folder="reports",
                public_id=random_public_id,
                overwrite=True,
                resource_type="image"
            )

            print("UPLOAD SUCCESS")

            cloudinary_url = upload_result.get("secure_url")

            print(cloudinary_url)

            Report.objects.create(
                store=store,
                item=item,
                image=cloudinary_url, # Lưu biến chứa link string vào trường image
                report_date=report_date
            )

            print("DB SAVE SUCCESS")

            return redirect("/dashboard/")

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse(
                f"<pre>{traceback.format_exc()}</pre>",
                status=500
            )

    return render(
        request,
        "checklist/upload.html",
        {"item": item}
    )
    
def admin_dashboard(request):

    if not request.user.is_staff:
        return redirect("/")    
    today = timezone.now().date()

    return render(
        request,
        "checklist/admin_dashboard.html",
        {
            "today": today
        }
    )
    
def admin_day(request, date):

    if not request.user.is_staff:
        return redirect("/")
    
    selected_date = datetime.strptime(
        date,
        "%Y-%m-%d"
    ).date()

    stores = Store.objects.all()

    total_stores = 0
    completed_stores = 0
    partial_stores = 0
    missing_stores = 0
    
    data = []

    total_items = ChecklistItem.objects.count()

    for store in stores:

        completed = Report.objects.filter(
            store=store,
            report_date=selected_date
        ).count()

        percent = 0

        if total_items > 0:
            percent = round(
                completed * 100 / total_items
            )

        data.append({
            "store": store,
            "completed": completed,
            "total": total_items,
            "percent": percent
        })
        
        if percent == 100:
            completed_stores += 1

        elif percent == 0:
            missing_stores += 1

        else:
            partial_stores += 1

        total_stores += 1

    system_percent = 0

    if total_stores > 0:

        system_percent = round(
            completed_stores * 100 / total_stores
        )
    
    return render(
        request,
        "checklist/admin_day.html",
        {
            "selected_date": selected_date,
            "data": data,
            "total_stores": total_stores,
            "completed_stores": completed_stores,
            "partial_stores": partial_stores,
            "missing_stores": missing_stores,
            "system_percent": system_percent,
        }
    )
    

def admin_store(request, store_id, date):

    if not request.user.is_staff:
        return redirect("/")

    selected_date = datetime.strptime(
        date,
        "%Y-%m-%d"
    ).date()

    store = get_object_or_404(
        Store,
        id=store_id
    )

    # =========================
    # AJAX
    # =========================
    if request.method == "POST":

        action = request.POST.get("action")

        # =========================
        # DUYỆT TẤT CẢ
        # pending -> pass
        # fail -> giữ nguyên
        # pass -> giữ nguyên
        # chưa chụp -> bỏ qua
        # =========================
        if action == "approve_all":

            reports = Report.objects.filter(
                store=store,
                report_date=selected_date
            )

            updated = reports.filter(
                status="pending"
            ).update(
                status="pass"
            )

            return JsonResponse({
                "success": True,
                "updated": updated
            })

        # =========================
        # ĐÁNH GIÁ TỪNG CHECKLIST
        # =========================
        report_id = request.POST.get("report_id")
        status = request.POST.get("status")
        note = request.POST.get("note")

        try:

            report = Report.objects.get(
                id=report_id,
                store=store,
                report_date=selected_date
            )

            report.status = status
            report.note = note
            report.save()

            return JsonResponse({
                "success": True
            })

        except Report.DoesNotExist:

            return JsonResponse({
                "success": False,
                "message": "Report not found"
            }, status=404)

    # =========================
    # LOAD PAGE
    # =========================

    slots = TimeSlot.objects.all().order_by(
        "start_time"
    )

    data = []

    for slot in slots:

        items = ChecklistItem.objects.filter(
            slot=slot
        ).order_by("id")

        row_items = []

        for item in items:

            report = Report.objects.filter(
                store=store,
                item=item,
                report_date=selected_date
            ).first()

            row_items.append({
                "item": item,
                "report": report
            })

        data.append({
            "slot": slot,
            "items": row_items
        })

    return render(
        request,
        "checklist/admin_store.html",
        {
            "store": store,
            "selected_date": selected_date,
            "data": data
        }
    )



def manage_checklist(request):

    if not request.user.is_staff:
        return redirect("/")
    if request.method == "POST":

        slot_id = request.POST.get("slot_id")
        title = request.POST.get("title")

        if title:

            ChecklistItem.objects.create(
                slot_id=slot_id,
                title=title
            )

        return redirect("/manage-checklist/")

    slots = TimeSlot.objects.all().order_by("start_time")

    return render(
        request,
        "checklist/manage_checklist.html",
        {
            "slots": slots
        }
    )
    

@require_POST
def delete_item(request, item_id):
    item = get_object_or_404(ChecklistItem, id=item_id)
    item.delete()
    return redirect("/manage-checklist/")
    
def admin_day_redirect(request):

    selected_date = request.GET.get(
        "date"
    )

    return redirect(
        f"/admin-day/{selected_date}/"
    )
    
def history(request):

    if "store_id" not in request.session:
        return redirect("/")

    store = Store.objects.get(
        id=request.session["store_id"]
    )

    selected_date = request.GET.get(
        "date"
    )

    slots = TimeSlot.objects.all()

    data = []

    for slot in slots:

        reports = Report.objects.filter(
            store=store,
            item__slot=slot
        )

        if selected_date:

            selected = parse_date(
                selected_date
            )

            if selected:

                reports = reports.filter(
                    report_date=parse_date(
                        selected_date
                    )
                )

        reports = reports.select_related(
            "item"
        ).order_by(
            "-created_at"
        )

        data.append({
            "slot": slot,
            "reports": reports
        })

    dates = Report.objects.filter(
        store=store
    ).dates(
        "created_at",
        "day",
        order="DESC"
    )

    return render(
        request,
        "checklist/history.html",
        {
            "store": store,
            "data": data,
            "dates": dates,
            "selected_date": selected_date
        }
    )
    
def export_excel(request):
    if not request.user.is_staff:
        return redirect("/")
    
    if request.method != "POST":

        return redirect(
            "/admin-dashboard/"
        )

    date_str = request.POST.get(
        "date"
    )

    selected_date = datetime.strptime(
        date_str,
        "%Y-%m-%d"
    ).date()

    store_ids = request.POST.getlist(
        "stores"
    )

    stores = Store.objects.filter(
        id__in=store_ids
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "QC Report"
    
    header_fill = PatternFill(
        "solid",
        fgColor="4472C4"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "Khung giờ",
        "Checklist"
    ]

    for store in stores:

        headers.append(
            f"CH {store.code}"
        )

    ws.append(headers)
    
    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center

        cell.border = thin

    green_fill = PatternFill(
        "solid",
        fgColor="92D050"
    )

    red_fill = PatternFill(
        "solid",
        fgColor="FF6666"
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFD966"
    )

    gray_fill = PatternFill(
        "solid",
        fgColor="D9D9D9"
    )

    current_row = 2

    for slot in TimeSlot.objects.all():

        items = ChecklistItem.objects.filter(
            slot=slot
        )

        slot_start_row = current_row

        first_item = True

        for item in items:

            row = []

            if first_item:

                row.append(
                    slot.name
                )

                first_item = False

            else:

                row.append("")

            row.append(
                item.title
            )

            for store in stores:

                report = Report.objects.filter(
                    store=store,
                    item=item,
                    report_date=selected_date
                ).first()

                if report:

                    if report.status == "pass":

                        value = "Đạt"

                    elif report.status == "fail":

                        value = "Không đạt"

                    else:

                        value = "Chưa đánh giá"

                else:

                    value = "Chưa báo cáo"

                row.append(value)

            ws.append(row)

            current_row += 1

        slot_end_row = current_row - 1

        if slot_end_row > slot_start_row:

            ws.merge_cells(
                start_row=slot_start_row,
                end_row=slot_end_row,
                start_column=1,
                end_column=1
            )

        slot_cell = ws.cell(
            row=slot_start_row,
            column=1
        )

        slot_cell.fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

        slot_cell.font = Font(
            bold=True
        )

        slot_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        slot_cell.border = thin

    for row in ws.iter_rows(
        min_row=2
    ):

        for cell in row:

            if cell.value == "Đạt":

                cell.fill = green_fill

            elif cell.value == "Không đạt":

                cell.fill = red_fill

            elif cell.value == "Chưa đánh giá":

                cell.fill = yellow_fill

            elif cell.value == "Chưa báo cáo":

                cell.fill = gray_fill

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="QC_{selected_date}.xlsx"'
    )
    
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40

    for col in range(
        3,
        ws.max_column + 1
    ):

        letter = ws.cell(
            row=1,
            column=col
        ).column_letter

        ws.column_dimensions[
            letter
        ].width = 18
    
    for row in ws.iter_rows():

        ws.row_dimensions[
            row[0].row
        ].height = 30
    
    for row in ws.iter_rows():

        for cell in row:

            cell.border = thin

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            
    ws.auto_filter.ref = ws.dimensions

    # ===============================
    # THỐNG KÊ
    # ===============================

    stat_start_row = ws.max_row + 2

    blue_fill = PatternFill("solid", fgColor="5B9BD5")
    orange_fill = PatternFill("solid", fgColor="F4B183")
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")

    # ===== ROW 1: TỶ LỆ ĐẠT =====
    ws.merge_cells(
        start_row=stat_start_row,
        start_column=1,
        end_row=stat_start_row,
        end_column=2
    )

    cell = ws.cell(row=stat_start_row, column=1)
    cell.value = "THỐNG KÊ TỶ LỆ ĐẠT"
    cell.fill = blue_fill
    cell.font = white_font
    cell.alignment = center

    # ===== ROW 2: LỖI =====
    ws.merge_cells(
        start_row=stat_start_row + 1,
        start_column=1,
        end_row=stat_start_row + 1,
        end_column=2
    )

    cell2 = ws.cell(row=stat_start_row + 1, column=1)
    cell2.value = "THỐNG KÊ LỖI / CHƯA BÁO CÁO"
    cell2.fill = orange_fill
    cell2.font = bold_font
    cell2.alignment = center

    # ===== FORMULA =====
    data_start = 2
    data_end = stat_start_row - 1

    for col in range(3, ws.max_column + 1):

        letter = ws.cell(row=1, column=col).column_letter

        # tỷ lệ đạt
        rate_formula = (
            f'=COUNTIF({letter}{data_start}:{letter}{data_end},"Đạt")'
            f'/COUNTA({letter}{data_start}:{letter}{data_end})'
        )

        c = ws.cell(row=stat_start_row, column=col)
        c.value = rate_formula
        c.number_format = "0.00%"
        c.fill = blue_fill
        c.font = white_font
        c.alignment = center

    # lỗi
        error_formula = (
            f'=COUNTA({letter}{data_start}:{letter}{data_end})'
            f'-COUNTIF({letter}{data_start}:{letter}{data_end},"Đạt")'
        )

        c2 = ws.cell(row=stat_start_row + 1, column=col)
        c2.value = error_formula
        c2.fill = orange_fill
        c2.font = bold_font
        c2.alignment = center

    wb.save(response)

    return response

def login_view(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        # ======================
        # ADMIN DJANGO
        # ======================
        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None and user.is_staff:

            login(request, user)

            today = timezone.localdate()

            return redirect(
                f"/admin-day/{today}/"
            )

        # ======================
        # STORE
        # ======================
        store = Store.objects.filter(
            code=username,
            password=password
        ).first()

        if store:

            request.session["store_id"] = store.id

            return redirect("/dashboard/")

        return render(
            request,
            "login.html",
            {
                "error": "Sai tài khoản hoặc mật khẩu"
            }
        )

    return render(
        request,
        "login.html"
    )
    
def audit_create(request):

    stores = Store.objects.all()

    if request.method == "POST":

        store_id = request.POST.get("store_id")

        audit = Audit.objects.create(
            store_id=store_id,
            user=request.user,
            score=200
        )

        for category in AuditCategory.objects.filter(
            is_active=True
        ):

            for item in category.items.filter(
                is_active=True
            ):

                status = request.POST.get(
                    f"status_{item.id}",
                    "pass"
                )

                note = request.POST.get(
                    f"note_{item.id}",
                    ""
                )

                issue = AuditIssue.objects.create(
                    audit=audit,
                    item=item,
                    category=item.category,
                    status=status,
                    note=note,
                    deduct_score=item.score if status == "fail" else 0
                )

                if status == "fail":

                    files = request.FILES.getlist(
                        f"image_{item.id}"
                    )

                    for file in files:

                        upload = cloudinary.uploader.upload(
                            file
                        )

                        AuditIssueImage.objects.create(
                            issue=issue,
                            image=upload["secure_url"]
                        )

        update_audit_score(audit)

        return redirect(
            f"/audit/{audit.id}/"
        )

    categories = AuditCategory.objects.prefetch_related(
        "items"
    ).all()

    return render(
        request,
        "checklist/audit_create.html",
        {
            "categories": categories,
            "stores": stores
        }
    )
    
def audit_detail(request, id):

    if not request.user.is_staff:
        return redirect("/")

    audit = get_object_or_404(Audit, id=id)

    # 👉 lấy issue đầu tiên theo status
    first_issue = audit.issues.order_by("status").first()

    return render(
        request,
        "checklist/audit_detail.html",
        {
            "audit": audit,
            "first_issue": first_issue
        }
    )
    
def audit_list(request):

    if not request.user.is_staff:
        return redirect("/")

    audits = Audit.objects.all().order_by("-created_at")

    search = request.GET.get("search")

    if search:
        audits = audits.filter(store__code__icontains=search)

    # ✔ CHECK có issue pending hay không
    audits = audits.annotate(

        is_pending=Exists(
            AuditIssue.objects.filter(
                audit_id=OuterRef("id"),
                status="pending"
            )
        ),

        has_fail=Exists(
            AuditIssue.objects.filter(
                audit_id=OuterRef("id"),
                status="fail"
            )
        ),

        need_review=Exists(
            AuditIssue.objects.filter(
                audit_id=OuterRef("id"),
                status="fixed"
            )
        )
    )



    paginator = Paginator(audits, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "checklist/audit_list.html", {
        "page_obj": page_obj,
        "search": search
    })
    
def delete_audit(request, id):

    audit = get_object_or_404(Audit, id=id)

    if request.method == "POST":
        audit.delete()
        messages.success(request, "Đã xoá audit thành công")
        return redirect("audit_list")

    return redirect("audit_list")

def staff_dashboard(request):

    if "store_id" not in request.session:
        return redirect("/")

    store = Store.objects.get(id=request.session["store_id"])

    audit_id = request.GET.get("audit_id")

    if audit_id:
        audit = get_object_or_404(Audit, id=audit_id, store=store)
    else:
        audit = Audit.objects.filter(store=store).order_by("-created_at").first()

    return render(request, "checklist/staff_dashboard.html", {
        "audit": audit
    })
    
def staff_fix_issue(request, id):

    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid method"})

    store_id = request.session.get("store_id")

    if not store_id:
        return JsonResponse({"success": False, "message": "No session"})

    store = Store.objects.filter(id=store_id).first()

    if not store:
        return JsonResponse({"success": False, "message": "Store not found"})

    issue = get_object_or_404(
        AuditIssue,
        id=id,
        audit__store=store
    )

    try:
        files = request.FILES.getlist("fix_image")

        if not files:
            return JsonResponse({"success": False, "message": "No files"})

        for file in files:
            upload = cloudinary.uploader.upload(file)

            AuditFixImage.objects.create(
                issue=issue,
                image=upload["secure_url"]
            )

        issue.status = "fixed"
        issue.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": str(e)
        })
        
def update_audit_score(audit):

    score = 200

    for issue in audit.issues.all():

        if issue.status == "fail":
            score -= issue.deduct_score

    audit.score = max(score, 0)
    audit.save()
    
def store_audit_history(request, store_id):

    audits = Audit.objects.filter(store_id=store_id).order_by("-created_at")

    return render(request, "checklist/store_history.html", {
        "audits": audits
    })
    
def staff_dashboard_by_audit(request, audit_id):

    store_id = request.session.get("store_id")

    if not store_id:
        return redirect("/")

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return redirect("/")

    audit = Audit.objects.filter(
        id=audit_id,
        store=store
    ).first()

    if not audit:
        return redirect(f"/store/{store.id}/audits/")

    return render(
        request,
        "checklist/staff_dashboard.html",
        {
            "audit": audit
        }
    )
    
def review_issue(request,audit_id,issue_id):

    issue=get_object_or_404(
        AuditIssue,
        id=issue_id
    )

    if request.method=="POST":

        status=request.POST.get("status")

        issue.status=status

        issue.save()

        update_audit_score(
            issue.audit
        )

    return redirect(
        f"/audit/{audit_id}/"
    )

def audit_review_issue(request, audit_id, issue_id):

    issue = get_object_or_404(
        AuditIssue,
        id=issue_id,
        audit_id=audit_id
    )

    if request.method == "POST":

        status = request.POST.get("status")

        issue.status = status
        issue.save()

    return redirect(
        f"/audit/{audit_id}/"
    )

def kpi_dashboard(request):

    if not request.user.is_staff:
        return redirect("/")

    stores = Store.objects.all().order_by("code")

    selected_stores = request.GET.getlist("stores")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    reports = Report.objects.all()

    if start_date:
        reports = reports.filter(
            report_date__gte=start_date
        )

    if end_date:
        reports = reports.filter(
            report_date__lte=end_date
        )

    if selected_stores:
        reports = reports.filter(
            store_id__in=selected_stores
        )

    total_reports = reports.count()

    pass_count = reports.filter(
        status="pass"
    ).count()

    fail_count = reports.filter(
        status="fail"
    ).count()

    pending_count = reports.filter(
        status="pending"
    ).count()

    if total_reports > 0:
        pass_percent = round(
            pass_count * 100 / total_reports,
            1
        )
    else:
        pass_percent = 0

    top_stores = (
        reports
        .values(
            "store__code"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total"
        )[:10]
    )

    top_fail_items = (
        reports.filter(
            status="fail"
        )
        .values(
            "item__title"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total"
        )[:10]
    )

    return render(
        request,
        "checklist/kpi_dashboard.html",
        {
            "stores": stores,

            "selected_stores": list(
                map(int, selected_stores)
            ) if selected_stores else [],

            "start_date": start_date,
            "end_date": end_date,

            "total_reports": total_reports,

            "pass_count": pass_count,

            "fail_count": fail_count,

            "pending_count": pending_count,

            "pass_percent": pass_percent,

            "top_stores": top_stores,

            "top_fail_items": top_fail_items
        }
    )

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

HEADER_FILL = PatternFill("solid", fgColor="E60012")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)

CENTER = Alignment(horizontal="center", vertical="center")

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

def export_kpi_excel(request):

    if not request.user.is_staff:
        return redirect("/")

    start_date=request.GET.get("start_date")
    end_date=request.GET.get("end_date")

    selected_stores = request.GET.getlist("stores")
    reports=Report.objects.all()
    
    if start_date:
        reports = reports.filter(
            report_date__gte=start_date
        )

    if end_date:
        reports = reports.filter(
            report_date__lte=end_date
        )

    stores = Store.objects.all()

    if selected_stores:
        stores = stores.filter(
            id__in=selected_stores
        )
        reports = reports.filter(
            store_id__in=selected_stores
        )

    display_start = "Tất cả"
    display_end = "Tất cả"

    if start_date:
        display_start = datetime.strptime(
            start_date,
            "%Y-%m-%d"
        ).strftime("%d/%m/%Y")

    if end_date:
        display_end = datetime.strptime(
            end_date,
            "%Y-%m-%d"
        ).strftime("%d/%m/%Y")

    wb=Workbook()

    ws = wb.active
    ws.title = "Tong Quan"

    ws.append(["Chỉ số", "Giá trị"])

    ws.append(["Từ ngày", display_start])
    ws.append(["Đến ngày", display_end])

    total = reports.count()

    passed = reports.filter(status="pass").count()
    failed = reports.filter(status="fail").count()
    pending = reports.filter(status="pending").count()

    percent = 0

    if total > 0:
        percent = round(passed * 100 / total, 1)

    ws.append(["Tổng checklist", total])
    ws.append(["Đạt", passed])
    ws.append(["Không đạt", failed])
    ws.append(["Pending", pending])
    ws.append(["Tỷ lệ đạt (%)", percent])

    format_sheet(ws)

    
    # SHEET 2
    ws2 = wb.create_sheet("Theo CH")

    ws2.append([
        "Cửa hàng",
        "Tổng lỗi/Chưa báo cáo",
        "Checklist đạt TB/ngày",
        "Tỷ lệ đat"
    ])

    # số ngày trong khoảng admin chọn
    days_count = reports.values(
        "report_date"
    ).distinct().count()

    for store in stores:

        rs = reports.filter(
            store=store
        )

        total_store = rs.count()

        # số checklist đạt
        pass_store = rs.filter(
            status="pass"
        ).count()

        # số lỗi
        fail_store = rs.filter(
            status__in=["fail", "pending"]
        ).count()

        # checklist đạt trung bình mỗi ngày
        avg_pass = 0

        if days_count > 0:

            avg_pass = round(
                pass_store / days_count,
                1
            )

        # % đạt của cửa hàng
        percent_store = 0

        if total_store > 0:

            percent_store = round(
                pass_store * 100 / total_store,
                1
            )

        ws2.append([
            store.code,
            fail_store,
            avg_pass,
            percent_store
        ])

    # tô màu cột % đạt
    for row in ws2.iter_rows(min_row=2):

        percent_cell = row[3]

        if percent_cell.value >= 95:

            percent_cell.fill = GREEN_FILL

        elif percent_cell.value >= 80:

            percent_cell.fill = YELLOW_FILL

        else:

            percent_cell.fill = RED_FILL

    format_sheet(ws2)
    
    #SHEET 3
    ws3 = wb.create_sheet("Top Loi")

    ws3.append([
        "CH",
        "Checklist lỗi",
        "Số lần"
    ])

    top_fail = (
        reports
        .filter(status="fail")
        .values(
            "store__code",
            "item__title"
        )
        .annotate(
            total=Count("id")
        )   
        .order_by("-total")
    )

    for row in top_fail:

        ws3.append([
            row["store__code"],
            row["item__title"],
            row["total"]
        ])

    format_sheet(ws3)
    
    #SHEET 4
    ws4 = wb.create_sheet("Theo Ngày")

    ws4.append([
        "Ngày",
        "% đạt trong ngày",
        "CH tệ nhất",
        "Tổng Lỗi/Chưa báo cáo trong ngày"
    ])

    # lấy danh sách ngày
    days = (
        reports
        .values_list("report_date", flat=True)
        .distinct()
        .order_by("report_date")
    )

    stores = list(stores)

    TOTAL_CHECKLIST = ChecklistItem.objects.count()

    for date_value in days:

        day_reports = reports.filter(report_date=date_value)

        # ==========================
        # % đạt toàn hệ thống trong ngày
        # ==========================
        pass_day = day_reports.filter(
            status="pass"
        ).count()

        expected = TOTAL_CHECKLIST * len(stores)

        percent_day = round(
            pass_day * 100 / expected,
            1
        ) if expected else 0

        # ==========================
        # Store tệ nhất
        # ==========================
        worst_store = None
        lowest_percent = 101

        for store in stores:

            store_rs = day_reports.filter(store=store)

            pass_store = store_rs.filter(status="pass").count()
            total_store = store_rs.count()

            percent_store = round((pass_store * 100 / TOTAL_CHECKLIST), 1) if TOTAL_CHECKLIST else 0

            if percent_store < lowest_percent:
                lowest_percent = percent_store
                worst_store = store.code

        # ==========================
        # Tổng lỗi + pending + chưa báo cáo
        # ==========================
        total_issue = 0

        for store in stores:

            store_rs = day_reports.filter(store=store)

            fail = store_rs.filter(status="fail").count()
            pending = store_rs.filter(status="pending").count()

            # số checklist chưa có report trong ngày
            reported_count = store_rs.count()
            not_report = max(TOTAL_CHECKLIST - reported_count, 0)

            total_issue += fail + pending + not_report

        ws4.append([
            date_value,
            percent_day,
            worst_store,
            f"{total_issue} lỗi"
        ])

    format_sheet(ws4)
    
    response=HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ]=(
        'attachment; filename="KetquaChecklist.xlsx"'
    )
    
    ws5 = wb.create_sheet("Leaderboard")

    red_fill=PatternFill(
        "solid",
        fgColor="E60012"
    )

    white_font=Font(
        bold=True,
        color="FFFFFF"
    )

    thin=Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
        
    ws5.append([
        "Hạng",
        "Cửa hàng",
        "% đạt",
        "Đã báo cáo",
        "Phải báo cáo"
    ])

    ranking=[]
    
    days_count = reports.values(
        "report_date"
    ).distinct().count()

    total_checklist = ChecklistItem.objects.count()
    
    for store in stores:

        rs=reports.filter(
            store=store
        )

        reported = rs.count()

        pass_store = rs.filter(
            status="pass"
        ).count()

        total_expected = total_checklist * days_count

        percent_store = 0

        if total_expected > 0:

            percent_store = round(
                pass_store * 100 / total_expected,
                1
            )

        ranking.append({
            "code": store.code,
            "percent": percent_store,
            "reported": reported,
            "expected": total_expected
        })
    
    ranking.sort(
        key=lambda x:x["percent"],
        reverse=True
    )
    
    rank = 1

    for row in ranking:

        ws5.append([
            rank,
            row["code"],
            row["percent"],
            row["reported"],
            row["expected"]
        ])

        current_row = ws5.max_row

        if rank == 1:

            fill = PatternFill(
                "solid",
                fgColor="FFD700"
            )

        elif rank == 2:

            fill = PatternFill(
                "solid",
                fgColor="C0C0C0"
            )

        elif rank == 3:

            fill = PatternFill(
                "solid",
                fgColor="CD7F32"
            )

        else:

            fill = None

        if fill:

            for cell in ws5[current_row]:

                cell.fill = fill

        rank += 1

    format_sheet(ws5)  
    
    auto_fit_columns(ws)
    auto_fit_columns(ws2)
    auto_fit_columns(ws3)
    auto_fit_columns(ws4)
    auto_fit_columns(ws5)
    
    wb.save(response)
    return response

def format_sheet(sheet):

    sheet.freeze_panes = "A2"

    # style từng ô
    for row in sheet.iter_rows():

        for cell in row:

            if isinstance(cell.value, (int, float)):

                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center"
                )

            else:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            cell.border = THIN

    # header
    for cell in sheet[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # filter
    sheet.auto_filter.ref = sheet.dimensions

    # auto width
    for col in sheet.columns:

        max_len = max(
            len(str(c.value))
            if c.value is not None
            else 0
            for c in col
        )

        sheet.column_dimensions[
            col[0].column_letter
        ].width = max_len + 5

    # format số
    for row in sheet.iter_rows(min_row=2):

        for cell in row:

            header = str(
                sheet.cell(
                    row=1,
                    column=cell.column
                ).value or ""
            ).lower()

            # chỉ format các cột % đạt
            if header == "% đạt":

                if isinstance(cell.value, (int, float)):

                    cell.number_format = "0.0"

            # checklist TB/ngày
            elif "tb/ngày" in header:

                if isinstance(cell.value, (int, float)):

                    cell.number_format = "0.0"
                    
def auto_fit_columns(sheet):

    for col in sheet.columns:

        max_length = 0
        column = col[0].column_letter

        for cell in col:

            if cell.value is None:
                continue

            value = str(cell.value)

            # xử lý % và số dài
            length = len(value)

            if length > max_length:
                max_length = length

        adjusted_width = max_length + 4

        # giới hạn cho đẹp UI
        if adjusted_width > 40:
            adjusted_width = 40

        sheet.column_dimensions[column].width = adjusted_width
        
def export_kpi_auditqc_excel(request):

    start = request.GET.get("start")
    end = request.GET.get("end")

    audits = Audit.objects.all()

    if start and end:
        audits = audits.filter(created_at__date__range=[start, end])

    avg_all = audits.aggregate(avg=Avg("score"))["avg"] or 0

    store_scores = (
        audits.values("store__id", "store__code")
        .annotate(avg_score=Avg("score"))
        .order_by("-avg_score")
    )

    issues = AuditIssue.objects.filter(audit__in=audits)

    issue_count = (
        issues.values("item__title")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    repeated_issues_raw = (
        issues.values("audit__store__code", "item__title")
        .annotate(total=Count("id"))
        .filter(total__gte=2)
        .order_by("audit__store__code")
    )

    grouped = {}
    for r in repeated_issues_raw:
        grouped.setdefault(r["audit__store__code"], []).append(r)

    # ================= STYLE =================
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")

    def style_row(ws, row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center

    wb = openpyxl.Workbook()

    # ================= SHEET 1 =================
    ws1 = wb.active
    ws1.title = "KPI Tổng"

    ws1.append(["Cửa hàng", "Điểm TB", "Rank"])

    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    rank = 1

    for s in store_scores:

        row = [s["store__code"], round(s["avg_score"], 2), rank]
        ws1.append(row)

        current_row = ws1[ws1.max_row]

        # border + center
        for c in current_row:
            c.border = thin_border
            c.alignment = center

        # 🎨 RANK COLOR
        if rank == 1:
            fill = PatternFill("solid", fgColor="FFD700")  # vàng
        elif rank == 2:
            fill = PatternFill("solid", fgColor="C0C0C0")  # bạc
        elif rank == 3:
            fill = PatternFill("solid", fgColor="CD7F32")  # đồng
        else:
            fill = None

        if fill:
            for c in current_row:
                c.fill = fill

        rank += 1

    # ================= SHEET 2 =================
    ws2 = wb.create_sheet("Lỗi phổ biến")

    ws2.append(["Rank", "Lỗi", "Số lần"])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center


    # CHỈ LẤY FAIL THỰC SỰ
    issue_count = (
        AuditIssue.objects
        .filter(
            audit__in=audits,
            status="fail"
        )
        .values("item__title")
        .annotate(total=Count("id"))
        .filter(total__gte=3)
        .order_by("-total")
    )

    rank = 1

    for i in issue_count:

        ws2.append([
            rank,
            i["item__title"],
            i["total"]
        ])

        row = ws2[ws2.max_row]
        total = i["total"]

        # style cơ bản
        for c in row:
            c.border = thin_border
            c.alignment = center

        # 🎨 màu theo mức độ
        if total >= 10:
            fill = PatternFill("solid", fgColor="FF4D4D")  # đỏ
            font = Font(color="FFFFFF", bold=True)

        elif total >= 5:
            fill = PatternFill("solid", fgColor="FFC000")  # vàng
            font = Font(color="000000", bold=True)

        else:
            fill = PatternFill("solid", fgColor="92D050")  # xanh
            font = Font(color="000000", bold=True)

        for c in row:
            c.fill = fill
            c.font = font

        rank += 1

    # ================= SHEET 3 =================
    ws3 = wb.create_sheet("Lỗi lặp lại")

    ws3.append(["Cửa hàng", "Lỗi", "Số lần"])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row_start = 2

    for store_code, items in grouped.items():

        start_row = row_start

        for i, item in enumerate(items):

            total = item["total"]

            ws3.append([
                store_code if i == 0 else "",
                item["item__title"],
                total
            ])

            # ===== STYLE CELL =====
            for col in range(1, 4):
                cell = ws3.cell(row=row_start, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # 🔴 nếu lỗi > 10 thì đỏ
                if col == 3:

                    if total > 10:
                        cell.fill = PatternFill("solid", fgColor="FF4D4D")  # đỏ
                        cell.font = Font(color="FFFFFF", bold=True)

                    elif total >= 5:
                        cell.fill = PatternFill("solid", fgColor="FFC000")  # vàng
                        cell.font = Font(color="000000", bold=True)

                    else:
                        cell.fill = PatternFill("solid", fgColor="92D050")  # xanh
                        cell.font = Font(color="000000", bold=True)

            row_start += 1

        end_row = row_start - 1

        # ===== MERGE SAFE =====
        if len(items) > 1 and start_row >= 2 and end_row >= start_row:

            ws3.merge_cells(
                start_row=start_row,
                end_row=end_row,
                start_column=1,
                end_column=1
            )

            ws3.cell(row=start_row, column=1).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            
    # ================= EXPORT =================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="AUDIT_QC_REPORT.xlsx"'

    wb.save(response)
    return response

def export_audit_excel(request, audit_id):
    audit = Audit.objects.select_related("store").get(id=audit_id)
    issues = audit.issues.select_related("item", "category").order_by("category__id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # Ensure grid lines are visible for a structured sheet look
    ws.views.sheetView[0].showGridLines = True

    # Premium Color Palette (Corporate Executive Theme)
    primary_navy = "1F3864"       # Deep Executive Navy for headers
    accent_blue = "2F5597"        # Secondary blue accent
    light_bg = "F9FBFD"           # Subtle cool background tint
    border_color = "D9D9D9"       # Soft grey border

    # Fonts
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color=primary_navy)
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    meta_label_font = Font(name=font_family, size=10, bold=True, color="595959")
    meta_val_font = Font(name=font_family, size=10, bold=True, color="262626")
    row_font = Font(name=font_family, size=10, color="000000")
    category_font = Font(name=font_family, size=10, bold=True, color=primary_navy)

    # Status Colors (Softer, modern tones)
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    green_font = Font(name=font_family, size=10, bold=True, color="375623")

    red_fill = PatternFill("solid", fgColor="FCE4D6")
    red_font = Font(name=font_family, size=10, bold=True, color="C65911")

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    yellow_font = Font(name=font_family, size=10, bold=True, color="806000")

    header_fill = PatternFill("solid", fgColor=primary_navy)
    meta_bg = PatternFill("solid", fgColor="F2F2F2")

    # Alignments & Borders
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_side = Side(style="thin", color=border_color)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.row_dimensions[1].height = 35
    ws["A1"] = "BÁO CÁO ĐÁNH GIÁ CHẤT LƯỢNG (QC AUDIT)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:E1")

    # Metadata Card styling (Store code & time)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    ws["A3"] = "MÃ CỬA HÀNG"
    ws["A3"].font = meta_label_font
    ws["A3"].fill = meta_bg
    ws["A3"].alignment = center
    ws["A3"].border = border

    ws["B3"] = audit.store.code
    ws["B3"].font = meta_val_font
    ws["B3"].alignment = center
    ws["B3"].border = border

    ws["A4"] = "THỜI GIAN CHẤM"
    ws["A4"].font = meta_label_font
    ws["A4"].fill = meta_bg
    ws["A4"].alignment = center
    ws["A4"].border = border

    local_time = timezone.localtime(audit.created_at).strftime("%d/%m/%Y %H:%M") if audit.created_at else "-"
    ws["B4"] = local_time
    ws["B4"].font = meta_val_font
    ws["B4"].alignment = center
    ws["B4"].border = border

    # Total Score KPI Box
    ws["A6"] = "TỔNG ĐIỂM QC"
    ws["A6"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws["A6"].fill = PatternFill("solid", fgColor=accent_blue)
    ws["A6"].alignment = center
    ws["A6"].border = border
    ws.row_dimensions[6].height = 26

    score_cell = ws["B6"]
    score_cell.border = border
    score_cell.alignment = center

    headers = ["Danh mục", "Hạng mục kiểm tra", "Trạng thái", "Điểm trừ", "Nội dung ghi chú / Chi tiết lỗi"]
    start_row = 8
    ws.row_dimensions[start_row].height = 26

    for col, h in enumerate(headers, 1):
        c = ws.cell(start_row, col, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    row = start_row + 1
    total_deduct = 0

    for category, group in groupby(issues, key=lambda x: x.category.title if x.category else "-"):
        category_start = row

        for issue in group:
            item = issue.item.title if issue.item else "-"
            score = getattr(issue, "deduct_score", 0) or 0
            total_deduct += score

            if issue.status == "pass":
                status = "ĐẠT"
                fill_cell = green_fill
                font_cell = green_font
            else:
                status = "KHÔNG ĐẠT"
                fill_cell = red_fill
                font_cell = red_font

            ws.row_dimensions[row].height = 22

            c1 = ws.cell(row, 1, category if row == category_start else None)
            c2 = ws.cell(row, 2, item)
            c3 = ws.cell(row, 3, status)
            c4 = ws.cell(row, 4, score if score > 0 else "-")
            c5 = ws.cell(row, 5, issue.note or "")

            c1.font = category_font
            c2.font = row_font
            c3.font = font_cell
            c4.font = row_font
            c5.font = row_font

            c3.fill = fill_cell

            for col_idx in range(1, 6):
                cell = ws.cell(row, col_idx)
                cell.border = border
                if col_idx in [3, 4]:
                    cell.alignment = center
                else:
                    cell.alignment = left

            row += 1

        category_end = row - 1
        if category_start != category_end:
            ws.merge_cells(start_row=category_start, start_column=1, end_row=category_end, end_column=1)
            ws.cell(category_start, 1).alignment = center

    final_score = 200 - total_deduct
    score_cell.value = final_score

    if final_score < 150:
        score_cell.fill = red_fill
        score_cell.font = Font(name=font_family, size=12, bold=True, color="C65911")
    elif final_score < 165:
        score_cell.fill = yellow_fill
        score_cell.font = Font(name=font_family, size=12, bold=True, color="806000")
    else:
        score_cell.fill = green_fill
        score_cell.font = Font(name=font_family, size=12, bold=True, color="375623")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 48

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="QC_Audit_{audit.store.code}.xlsx"'
    wb.save(response)

    return response